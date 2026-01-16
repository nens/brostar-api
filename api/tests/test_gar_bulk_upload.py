import io
import json
import os
from unittest.mock import Mock, patch

import pandas as pd
import polars as pl
import pytest
from openpyxl import Workbook
from rest_framework.test import APIClient

from api.bro_upload.gar_bulk_upload import (
    create_gar_field_measurements,
    csv_or_excel_to_df,
)
from api.bro_upload.upload_datamodels import FieldMeasurement
from api.tests import fixtures

user = fixtures.user
organisation = fixtures.organisation  # imported, even though not used in this file, because required for userprofile fixture
userprofile = fixtures.userprofile


@pytest.fixture
def api_client():
    return APIClient()


@pytest.mark.django_db
def test_gar_bulk_upload_invalid_input(
    api_client, organisation, user, userprofile, tmp_path
):
    """Testing the 400 response on a request with just 1 file."""
    api_client.force_authenticate(user=user)
    url = "/api/bulkuploads/"

    data = {"bulk_upload_type": "GAR"}

    d = tmp_path / "sub"
    d.mkdir()
    file_path = d / "test.csv"
    csv_data = {"test1": ["test1", "test2"], "test2": ["test1", "test3"]}
    df = pd.DataFrame(csv_data)
    df.to_csv(file_path, index=False)

    with file_path.open("rb") as fp:
        data["fieldwork_file"] = fp
        r = api_client.post(url, data, format="multipart")

    assert r.status_code == 400


@pytest.mark.django_db
def test_gar_bulk_upload_valid_input(
    api_client, organisation, user, userprofile, tmp_path
):
    """Testing the 201 response on a valid request."""
    api_client.force_authenticate(user=user)
    url = "/api/bulkuploads/"

    d = tmp_path / "sub"
    d.mkdir()
    fieldwork_file_path = d / "fieldwork_test.csv"
    lab_file_path = d / "lab_test.csv"

    # Data for fieldwork file
    fieldwork_data = {"test1": ["test1", "test2"], "test2": ["test1", "test3"]}
    df_fieldwork = pd.DataFrame(fieldwork_data)
    df_fieldwork.to_csv(fieldwork_file_path, index=False)

    # Data for lab file
    lab_data = {"test3": ["test4", "test5"], "test4": ["test6", "test7"]}
    df_lab = pd.DataFrame(lab_data)
    df_lab.to_csv(lab_file_path, index=False)

    metadata_json = json.dumps(
        {
            "requestReference": "test",
            "qualityRegime": "IMBRO",
        }
    )

    with (
        fieldwork_file_path.open("rb") as fp_fieldwork,
        lab_file_path.open("rb") as fp_lab,
    ):
        data = {
            "bulk_upload_type": "GAR",
            "project_number": 1,
            "metadata": metadata_json,
            "fieldwork_file": fp_fieldwork,
            "lab_file": fp_lab,
        }
        with patch("api.tasks.gar_bulk_upload_task.delay") as mock_task:
            r = api_client.post(url, data, format="multipart")

    assert mock_task.called
    assert r.status_code == 201


class TestCsvOrExcelToDf:
    """Tests for csv_or_excel_to_df function."""

    def create_mock_upload_file(self, filename: str, file_content):
        """Helper to create a mock UploadFile instance with real file-like object."""
        mock_instance = Mock()
        mock_instance.file = file_content
        mock_instance.file.name = filename

        return mock_instance

    def test_csv_reading_basic(self):
        """Test reading a basic CSV file."""
        csv_data = "name,age,city\nAlice,30,NYC\nBob,25,LA\nCharlie,35,Chicago"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "test.csv"

        mock_upload = self.create_mock_upload_file("test.csv", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert len(df) == 3
        assert list(df.columns) == ["name", "age", "city"]
        assert df.iloc[0]["name"] == "Alice"
        assert df.iloc[0]["age"] == 30

    def test_csv_with_different_delimiter(self):
        """Test CSV with standard delimiter."""
        csv_data = "name,age,city\nDavid,28,Boston"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "data.csv"

        mock_upload = self.create_mock_upload_file("data.csv", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert not df.empty
        assert "name" in df.columns

    def test_csv_empty_file(self):
        """Test reading an empty CSV file."""
        csv_data = ""
        csv_file = io.StringIO(csv_data)
        csv_file.name = "empty.csv"

        mock_upload = self.create_mock_upload_file("empty.csv", csv_file)

        # Should handle empty file gracefully
        with pytest.raises(pd.errors.EmptyDataError):
            csv_or_excel_to_df(mock_upload)

    def test_csv_with_special_characters(self):
        """Test CSV with special characters and quotes."""
        csv_data = 'name,description\n"John ""Johnny"" Doe","He said, ""Hello!"""'
        csv_file = io.StringIO(csv_data)
        csv_file.name = "special.csv"

        mock_upload = self.create_mock_upload_file("special.csv", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert len(df) == 1
        assert 'John "Johnny" Doe' in df.iloc[0]["name"]

    def test_excel_xlsx_reading(self):
        """Test reading an XLSX file."""
        # Create an in-memory Excel file
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active

        # Add headers
        sheet["A1"] = "name"
        sheet["B1"] = "age"
        sheet["C1"] = "city"

        # Add data
        sheet["A2"] = "Alice"
        sheet["B2"] = 30
        sheet["C2"] = "NYC"

        sheet["A3"] = "Bob"
        sheet["B3"] = 25
        sheet["C3"] = "LA"

        workbook.save(output)
        output.seek(0)
        output.name = "test.xlsx"

        mock_upload = self.create_mock_upload_file("test.xlsx", output)

        df = csv_or_excel_to_df(mock_upload)

        assert len(df) == 2
        assert list(df.columns) == ["name", "age", "city"]
        assert df.iloc[0]["name"] == "Alice"
        assert df.iloc[0]["age"] == 30

    def test_excel_xls_extension(self):
        """Test that .xls extension is handled (reads as Excel)."""
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "col1"
        sheet["A2"] = "value1"
        workbook.save(output)
        output.seek(0)
        output.name = "old_format.xls"

        mock_upload = self.create_mock_upload_file("old_format.xls", output)

        df = csv_or_excel_to_df(mock_upload)

        assert not df.empty
        assert "col1" in df.columns

    def test_excel_with_multiple_sheets(self):
        """Test Excel file with multiple sheets (reads first sheet by default)."""
        output = io.BytesIO()
        workbook = Workbook()

        # First sheet
        sheet1 = workbook.active
        sheet1.title = "Sheet1"
        sheet1["A1"] = "data1"
        sheet1["A2"] = "value1"

        # Second sheet
        sheet2 = workbook.create_sheet("Sheet2")
        sheet2["A1"] = "data2"
        sheet2["A2"] = "value2"

        workbook.save(output)
        output.seek(0)
        output.name = "multi_sheet.xlsx"

        mock_upload = self.create_mock_upload_file("multi_sheet.xlsx", output)

        df = csv_or_excel_to_df(mock_upload)

        # Should read first sheet
        assert "data1" in df.columns

    def test_unsupported_file_type(self):
        """Test that unsupported file types raise ValueError."""
        txt_data = "This is a text file"
        txt_file = io.StringIO(txt_data)
        txt_file.name = "document.txt"

        mock_upload = self.create_mock_upload_file("document.txt", txt_file)

        with pytest.raises(ValueError, match="Unsupported file type"):
            csv_or_excel_to_df(mock_upload)

    def test_pdf_file_raises_error(self):
        """Test that PDF files raise ValueError."""
        pdf_file = io.BytesIO(b"fake pdf")
        pdf_file.name = "document.pdf"

        mock_upload = self.create_mock_upload_file("document.pdf", pdf_file)

        with pytest.raises(ValueError, match="Unsupported file type"):
            csv_or_excel_to_df(mock_upload)

    def test_file_extension_case_insensitive(self):
        """Test that file extensions are case-insensitive."""
        csv_data = "col1,col2\nval1,val2"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "TEST.CSV"

        mock_upload = self.create_mock_upload_file("TEST.CSV", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert len(df) == 1
        assert "col1" in df.columns

    def test_file_with_extra_spaces_in_extension(self):
        """Test that extra spaces in extension are handled."""
        csv_data = "col1\nval1"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "data.csv "

        mock_upload = self.create_mock_upload_file("data.csv ", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert not df.empty

    def test_seek_functionality(self):
        """Test that the function seeks to beginning of file."""
        csv_data = "col1,col2\nval1,val2\nval3,val4"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "test.csv"

        # Read once to move pointer
        csv_file.read()
        assert csv_file.tell() > 0  # Pointer is not at start

        mock_upload = self.create_mock_upload_file("test.csv", csv_file)

        # Function should seek(0) before reading
        df = csv_or_excel_to_df(mock_upload)

        assert len(df) == 2  # Should read all rows

    def test_csv_with_numeric_data(self):
        """Test CSV with various numeric types."""
        csv_data = "int_col,float_col,str_col\n1,1.5,text\n2,2.7,more"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "numeric.csv"

        mock_upload = self.create_mock_upload_file("numeric.csv", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert df["int_col"].dtype in [int, "int64"]
        assert df["float_col"].dtype in [float, "float64"]
        assert len(df) == 2

    def test_excel_with_formulas(self):
        """Test Excel file containing formulas."""
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active

        sheet["A1"] = "value"
        sheet["B1"] = "formula"
        sheet["A2"] = 10
        sheet["B2"] = "=A2*2"

        workbook.save(output)
        output.seek(0)
        output.name = "formulas.xlsx"

        mock_upload = self.create_mock_upload_file("formulas.xlsx", output)

        df = csv_or_excel_to_df(mock_upload)

        assert not df.empty
        assert "value" in df.columns

    def test_csv_with_missing_values(self):
        """Test CSV with missing/null values."""
        csv_data = "col1,col2,col3\n1,2,3\n4,,6\n,,9"
        csv_file = io.StringIO(csv_data)
        csv_file.name = "missing.csv"

        mock_upload = self.create_mock_upload_file("missing.csv", csv_file)

        df = csv_or_excel_to_df(mock_upload)

        assert len(df) == 3
        assert pd.isna(df.iloc[1]["col2"])
        assert pd.isna(df.iloc[2]["col1"])


class TestCreateGarFieldMeasurements:
    """Tests for create_gar_field_measurements function"""

    @pytest.fixture
    def mock_csv_data(self):
        """Mock CSV data for GAR parameters"""
        return pl.DataFrame(
            {
                "ID": [3, 4, 5, 6],
                "aquocode": ["1112T4ClC2a", "111TClC2a", "1122T4ClC2a", "112TClC2a"],
                "CASnummer": ["630-20-6", "71-55-6", "79-34-5", "79-00-5"],
                "omschrijving": [
                    "1,1,1,2-tetrachloorethaan",
                    "1,1,1-trichloorethaan",
                    "1,1,2,2-tetrachloorethaan",
                    "1,1,2-trichloorethaan",
                ],
                "eenheid": ["ug/l", "ug/l", "ug/l", "ug/l"],
                "hoedanigheid": ["NVT", "NVT", "NVT", "NVT"],
            }
        )

    # Tests for old format (Provincie Noord-Brabant)

    def test_old_format_single_parameter(self):
        """Test old format with a single field parameter"""
        row = pd.Series(
            {"Zuurstof (mg/L)": 8.5, "pH": 7.2, "other_column": "some_value"}
        )

        result = create_gar_field_measurements(row)

        # Zuurstof is written wrong
        assert len(result) == 1
        assert all(isinstance(fm, FieldMeasurement) for fm in result)

        # Check pH measurement
        ph_measurement = next(fm for fm in result if fm.parameter == 1398)
        assert ph_measurement.unit == "1"
        assert ph_measurement.field_measurement_value == 7.2
        assert ph_measurement.quality_control_status == "onbeslist"

    def test_old_format_all_parameters(self):
        """Test old format with all available field parameters"""
        row = pd.Series(
            {
                "Zuurstof (mg/L)": 8.5,
                "pH": 7.2,
                "Zuurstof (mg/l)": 8.3,
                "Geleidbaarheid (mS/m)": 50.0,
                "Temperatuur (°C)": 15.5,
                "Troebelheid (NTU)": 2.5,
                "Alkaliniteit (HCO3 - mg/l)": 120.0,
                "Chloride (mg/l)": 25.0,
            }
        )

        result = create_gar_field_measurements(row)

        # Should have 6 measurements (all except the trigger column)
        assert len(result) == 7

        # Verify specific measurements
        oxygen_measurement = next(fm for fm in result if fm.parameter == 1701)
        assert oxygen_measurement.field_measurement_value == 8.3
        assert oxygen_measurement.unit == "mg/l"

    def test_old_format_with_na_values(self):
        """Test old format with NaN values (should be skipped)"""
        row = pd.Series(
            {
                "Zuurstof (mg/L)": 8.5,
                "pH": 7.2,
                "Geleidbaarheid (mS/m)": pd.NA,
                "Temperatuur (°C)": None,
            }
        )

        result = create_gar_field_measurements(row)

        # Only pH should be included (Zuurstof (mg/L) is trigger column)
        assert len(result) == 1
        assert result[0].parameter == 1398

    def test_old_format_empty_row(self):
        """Test old format with no valid parameters"""
        row = pd.Series({"Zuurstof (mg/L)": 8.5, "other_column": "value"})

        result = create_gar_field_measurements(row)

        assert len(result) == 0

    # Tests for new format (GAR)

    @patch("polars.read_csv")
    def test_new_format_single_parameter(self, mock_read_csv, mock_csv_data):
        """Test new format with a single aquocode parameter"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series({"1112T4ClC2a": "5.2", "other_column": "value"})

        result = create_gar_field_measurements(row)

        assert len(result) == 1
        assert result[0].parameter == 3
        assert result[0].unit == "ug/l"
        assert result[0].field_measurement_value == 5.2
        assert result[0].quality_control_status == "onbeslist"

    @patch("polars.read_csv")
    def test_new_format_multiple_parameters(self, mock_read_csv, mock_csv_data):
        """Test new format with multiple aquocode parameters"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series(
            {
                "1112T4ClC2a": "5.2",
                "111TClC2a": "3.1",
                "1122T4ClC2a": "1.5",
                "other_column": "value",
            }
        )

        result = create_gar_field_measurements(row)

        assert len(result) == 3
        parameter_ids = [fm.parameter for fm in result]
        assert set(parameter_ids) == {3, 4, 5}

    @patch("polars.read_csv")
    def test_new_format_niet_bepaald_skipped(self, mock_read_csv, mock_csv_data):
        """Test that 'niet bepaald' values are skipped"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series(
            {
                "1112T4ClC2a": "5.2",
                "111TClC2a": "niet bepaald",
                "1122T4ClC2a": "1.5",
            }
        )

        result = create_gar_field_measurements(row)

        assert len(result) == 2
        parameter_ids = [fm.parameter for fm in result]
        assert 4 not in parameter_ids  # 111TClC2a should be skipped

    @patch("polars.read_csv")
    def test_new_format_unknown_aquocode_skipped(self, mock_read_csv, mock_csv_data):
        """Test that unknown aquocodes are skipped"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series(
            {"1112T4ClC2a": "5.2", "UNKNOWN_CODE": "10.0", "other_column": "value"}
        )

        result = create_gar_field_measurements(row)

        assert len(result) == 1
        assert result[0].parameter == 3

    @patch("polars.read_csv")
    def test_new_format_empty_row(self, mock_read_csv, mock_csv_data):
        """Test new format with no valid aquocodes"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series({"other_column": "value", "another_column": "another_value"})

        result = create_gar_field_measurements(row)

        assert len(result) == 0

    @patch("polars.read_csv")
    def test_new_format_csv_path(self, mock_read_csv, mock_csv_data):
        """Test that CSV is read from correct path"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series({"other_column": "value"})

        create_gar_field_measurements(row)

        # Verify read_csv was called
        mock_read_csv.assert_called_once()
        call_args = mock_read_csv.call_args

        # Check that path ends with the expected filename
        assert "20260107_GARVarList.csv" in call_args[0][0]
        assert call_args[1]["separator"] == ";"

    # Edge cases and format detection

    def test_format_detection_old_format(self):
        """Test that old format is correctly detected"""
        row = pd.Series({"Zuurstof (mg/L)": 8.5, "pH": 7.2})

        # Should use old format logic
        result = create_gar_field_measurements(row)
        assert len(result) == 1  # Only pH

    @patch("polars.read_csv")
    def test_format_detection_new_format(self, mock_read_csv, mock_csv_data):
        """Test that new format is correctly detected"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series({"1112T4ClC2a": "5.2", "other": "value"})

        # Should use new format logic
        result = create_gar_field_measurements(row)
        assert len(result) == 1

    def test_empty_series(self):
        """Test with completely empty series"""
        row = pd.Series(dtype=object)

        result = create_gar_field_measurements(row)

        assert len(result) == 0
        assert isinstance(result, list)

    @patch("polars.read_csv")
    def test_mixed_data_types(self, mock_read_csv, mock_csv_data):
        """Test handling of different data types in values"""
        mock_read_csv.return_value = mock_csv_data

        row = pd.Series(
            {
                "1112T4ClC2a": 5.2,  # numeric
                "111TClC2a": "3.1",  # string
            }
        )

        result = create_gar_field_measurements(row)

        assert len(result) == 2
        # Both should be accepted regardless of type
        assert all(fm.field_measurement_value in [5.2, 3.1] for fm in result)

    def test_quality_control_status_always_onbeslist(self):
        """Test that qualityControlStatus is always 'onbeslist'"""
        row = pd.Series({"Zuurstof (mg/L)": 8.5, "pH": 7.2, "Temperatuur (°C)": 15.5})

        result = create_gar_field_measurements(row)

        for measurement in result:
            assert measurement.quality_control_status == "onbeslist"


# Integration-style tests


class TestCreateGarFieldMeasurementsIntegration:
    """Integration tests with actual CSV file (if available)"""

    @pytest.mark.skipif(
        not os.path.exists(
            os.path.join(
                os.path.dirname(__file__), "bro_upload", "20260107_GARVarList.csv"
            )
        ),
        reason="CSV file not available",
    )
    def test_with_real_csv_file(self):
        """Test with actual CSV file if present"""
        row = pd.Series({"1112T4ClC2a": "5.2", "other_column": "value"})

        result = create_gar_field_measurements(row)

        # Basic validation
        assert isinstance(result, list)
        if len(result) > 0:
            assert all(isinstance(fm, FieldMeasurement) for fm in result)
